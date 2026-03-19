from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_login import login_user, logout_user, login_required, current_user
from app import db, bcrypt, socketio
from app.models import User, Course, Enrollment, Attendance
from datetime import datetime, date, timedelta
from functools import wraps
import csv
import io
import json

main = Blueprint('main', __name__)


# ─── Decorators ────────────────────────────────────────────────
def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('main.login'))
            if current_user.role not in roles:
                flash('Access denied.', 'danger')
                return redirect(url_for('main.dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator


# ─── Auth Routes ───────────────────────────────────────────────
@main.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return render_template('index.html')


@main.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember') == 'on'

        user = User.query.filter(
            (User.username == username) | (User.email == username)
        ).first()

        if not user:
            flash('Invalid username or password.', 'danger')
            return render_template('login.html')

        # Check account lockout
        if user.locked_until and user.locked_until > datetime.utcnow():
            flash(f'Account locked. Try again after {user.locked_until.strftime("%H:%M")}.', 'danger')
            return render_template('login.html')

        if user.check_password(password) and user.is_active:
            user.failed_logins = 0
            user.locked_until = None
            db.session.commit()
            login_user(user, remember=remember)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('main.dashboard'))
        else:
            user.failed_logins = (user.failed_logins or 0) + 1
            if user.failed_logins >= 5:
                user.locked_until = datetime.utcnow() + timedelta(minutes=30)
                flash('Too many failed attempts. Account locked for 30 minutes.', 'danger')
            else:
                flash(f'Invalid password. {5 - user.failed_logins} attempts remaining.', 'danger')
            db.session.commit()

    return render_template('login.html')


@main.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('main.login'))


# ─── Dashboard Router ──────────────────────────────────────────
@main.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'admin':
        return redirect(url_for('main.admin_dashboard'))
    elif current_user.role == 'teacher':
        return redirect(url_for('main.teacher_dashboard'))
    else:
        return redirect(url_for('main.student_dashboard'))


# ─── Admin Routes ──────────────────────────────────────────────
@main.route('/admin')
@login_required
@role_required('admin')
def admin_dashboard():
    users = User.query.all()
    courses = Course.query.all()
    students = [u for u in users if u.role == 'student']
    teachers = [u for u in users if u.role == 'teacher']

    stats = {
        'total_students': len(students),
        'total_teachers': len(teachers),
        'total_courses': len(courses),
        'total_attendance': Attendance.query.count()
    }

    return render_template('admin_dashboard.html', users=users, courses=courses,
                           stats=stats, students=students, teachers=teachers)


@main.route('/admin/users/add', methods=['POST'])
@login_required
@role_required('admin')
def add_user():
    data = request.form
    if User.query.filter_by(username=data['username']).first():
        flash('Username already exists.', 'danger')
        return redirect(url_for('main.admin_dashboard'))
    if User.query.filter_by(email=data['email']).first():
        flash('Email already exists.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    user = User(
        username=data['username'],
        email=data['email'],
        role=data['role'],
        roll_no=data.get('roll_no') or None,
        parent_phone=data.get('parent_phone') or None,
        parent_email=data.get('parent_email') or None
    )
    user.set_password(data['password'])
    db.session.add(user)
    db.session.commit()
    flash(f'User {user.username} created successfully.', 'success')
    return redirect(url_for('main.admin_dashboard'))


@main.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@login_required
@role_required('admin')
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('Cannot delete yourself.', 'danger')
        return redirect(url_for('main.admin_dashboard'))
    db.session.delete(user)
    db.session.commit()
    flash(f'User {user.username} deleted.', 'success')
    return redirect(url_for('main.admin_dashboard'))


@main.route('/admin/users/bulk-upload', methods=['POST'])
@login_required
@role_required('admin')
def bulk_upload_users():
    file = request.files.get('csv_file')
    if not file:
        flash('No file uploaded.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    stream = io.StringIO(file.stream.read().decode('UTF8'))
    reader = csv.DictReader(stream)
    added, failed = 0, []

    for i, row in enumerate(reader, 2):
        try:
            if User.query.filter_by(username=row['username']).first():
                failed.append(f"Row {i}: Username '{row['username']}' exists")
                continue
            if User.query.filter_by(email=row['email']).first():
                failed.append(f"Row {i}: Email '{row['email']}' exists")
                continue
            user = User(
                username=row['username'].strip(),
                email=row['email'].strip(),
                role=row.get('role', 'student').strip(),
                roll_no=row.get('roll_no', '').strip() or None,
                parent_phone=row.get('parent_phone', '').strip() or None,
                parent_email=row.get('parent_email', '').strip() or None
            )
            user.set_password(row.get('password', 'Pass@1234'))
            db.session.add(user)
            added += 1
        except Exception as e:
            failed.append(f"Row {i}: {str(e)}")

    db.session.commit()
    flash(f'Bulk upload: {added} users added, {len(failed)} failed.', 'success' if not failed else 'warning')
    if failed:
        session['bulk_errors'] = failed[:10]
    return redirect(url_for('main.admin_dashboard'))


@main.route('/admin/courses/add', methods=['POST'])
@login_required
@role_required('admin')
def add_course():
    data = request.form
    if Course.query.filter_by(code=data['code']).first():
        flash('Course code already exists.', 'danger')
        return redirect(url_for('main.admin_dashboard'))
    course = Course(
        name=data['name'],
        code=data['code'].upper(),
        teacher_id=int(data['teacher_id'])
    )
    db.session.add(course)
    db.session.commit()
    flash(f'Course {course.code} created.', 'success')
    return redirect(url_for('main.admin_dashboard'))


@main.route('/admin/courses/delete/<int:course_id>', methods=['POST'])
@login_required
@role_required('admin')
def delete_course(course_id):
    course = Course.query.get_or_404(course_id)
    db.session.delete(course)
    db.session.commit()
    flash(f'Course {course.code} deleted.', 'success')
    return redirect(url_for('main.admin_dashboard'))


@main.route('/admin/analytics')
@login_required
@role_required('admin')
def admin_analytics():
    students = User.query.filter_by(role='student').all()
    courses = Course.query.all()
    teachers = User.query.filter_by(role='teacher').all()

    risk_counts = {'SAFE': 0, 'CAUTION': 0, 'WARNING': 0, 'CRITICAL': 0}
    student_stats = []
    for s in students:
        pct = s.get_attendance_percentage()
        risk, _ = s.get_risk_level()
        risk_counts[risk] += 1
        student_stats.append({'student': s, 'pct': pct, 'risk': risk})

    overall_avg = sum(x['pct'] for x in student_stats) / len(student_stats) if student_stats else 0

    course_stats = []
    for c in courses:
        enrolled = len(list(c.enrollments))
        avg = c.get_avg_attendance()
        at_risk = sum(1 for s in c.get_enrolled_students()
                      if s.get_attendance_percentage(c.id) < 75)
        course_stats.append({'course': c, 'enrolled': enrolled, 'avg': avg, 'at_risk': at_risk})

    return render_template('admin_analytics.html',
                           risk_counts=risk_counts, student_stats=student_stats,
                           overall_avg=overall_avg, course_stats=course_stats,
                           teachers=teachers)


@main.route('/admin/csv-template')
@login_required
@role_required('admin')
def csv_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['username', 'email', 'password', 'role', 'roll_no', 'parent_phone', 'parent_email'])
    writer.writerow(['john_doe', 'john@example.com', 'Pass@1234', 'student', 'CS101', '+919876543210', 'parent@example.com'])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='users_template.csv')


# ─── Teacher Routes ────────────────────────────────────────────
@main.route('/teacher')
@login_required
@role_required('teacher')
def teacher_dashboard():
    courses = Course.query.filter_by(teacher_id=current_user.id).all()
    course_data = []
    for c in courses:
        enrolled = len(list(c.enrollments))
        avg = c.get_avg_attendance()
        total_classes = c.get_total_classes()
        course_data.append({'course': c, 'enrolled': enrolled, 'avg': avg, 'total_classes': total_classes})
    return render_template('teacher_dashboard.html', course_data=course_data)


@main.route('/teacher/course/<int:course_id>/enrollments', methods=['GET', 'POST'])
@login_required
@role_required('teacher')
def manage_enrollments(course_id):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    if request.method == 'POST':
        action = request.form.get('action')
        student_ids = request.form.getlist('student_ids')

        if action == 'enroll':
            for sid in student_ids:
                if not Enrollment.query.filter_by(student_id=sid, course_id=course_id).first():
                    db.session.add(Enrollment(student_id=int(sid), course_id=course_id))
            db.session.commit()
            flash(f'{len(student_ids)} students enrolled.', 'success')
        elif action == 'unenroll':
            for sid in student_ids:
                e = Enrollment.query.filter_by(student_id=sid, course_id=course_id).first()
                if e:
                    db.session.delete(e)
            db.session.commit()
            flash(f'{len(student_ids)} students unenrolled.', 'success')

    enrolled_ids = [e.student_id for e in course.enrollments]
    enrolled_students = User.query.filter(User.id.in_(enrolled_ids)).all()
    all_students = User.query.filter_by(role='student').all()
    unenrolled = [s for s in all_students if s.id not in enrolled_ids]

    return render_template('manage_enrollments.html',
                           course=course, enrolled_students=enrolled_students,
                           unenrolled_students=unenrolled)


@main.route('/teacher/course/<int:course_id>/attendance', methods=['GET', 'POST'])
@login_required
@role_required('teacher')
def mark_attendance(course_id):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    selected_date = request.args.get('date', date.today().isoformat())
    try:
        att_date = date.fromisoformat(selected_date)
    except ValueError:
        att_date = date.today()

    if request.method == 'POST':
        att_date_str = request.form.get('date', date.today().isoformat())
        att_date = date.fromisoformat(att_date_str)

        enrolled_ids = [e.student_id for e in course.enrollments]
        updated = 0
        for sid in enrolled_ids:
            status = request.form.get(f'status_{sid}', 'absent')
            existing = Attendance.query.filter_by(
                student_id=sid, course_id=course_id, date=att_date).first()
            if existing:
                existing.status = status
                existing.marked_at = datetime.utcnow()
            else:
                db.session.add(Attendance(
                    student_id=sid, course_id=course_id,
                    date=att_date, status=status
                ))
            updated += 1

        db.session.commit()

        # Emit real-time update
        socketio.emit('attendance_updated', {
            'course_id': course_id,
            'date': att_date_str,
            'count': updated
        })

        flash(f'Attendance saved for {updated} students on {att_date}.', 'success')
        return redirect(url_for('main.mark_attendance', course_id=course_id, date=att_date_str))

    enrolled_ids = [e.student_id for e in course.enrollments]
    students = User.query.filter(User.id.in_(enrolled_ids)).order_by(User.roll_no).all()

    existing_att = {
        a.student_id: a.status
        for a in Attendance.query.filter_by(course_id=course_id, date=att_date).all()
    }

    return render_template('mark_attendance.html',
                           course=course, students=students,
                           existing_att=existing_att, selected_date=att_date)


@main.route('/teacher/course/<int:course_id>/analytics')
@login_required
@role_required('teacher')
def course_analytics(course_id):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    students = course.get_enrolled_students()
    total_classes = course.get_total_classes()

    student_analytics = []
    risk_counts = {'SAFE': 0, 'CAUTION': 0, 'WARNING': 0, 'CRITICAL': 0}

    for s in students:
        att_q = Attendance.query.filter_by(student_id=s.id, course_id=course_id)
        total = att_q.count()
        present = att_q.filter(Attendance.status.in_(['present', 'late'])).count()
        absent = total - present
        pct = s.get_attendance_percentage(course_id)
        risk, color = s.get_risk_level(course_id)
        can_miss, classes_needed = s.get_can_miss(course_id)
        risk_counts[risk] += 1

        student_analytics.append({
            'student': s, 'total': total, 'present': present,
            'absent': absent, 'pct': pct, 'risk': risk,
            'color': color, 'can_miss': can_miss, 'classes_needed': classes_needed
        })

    avg_pct = sum(x['pct'] for x in student_analytics) / len(student_analytics) if student_analytics else 0

    return render_template('course_analytics.html',
                           course=course, student_analytics=student_analytics,
                           risk_counts=risk_counts, avg_pct=avg_pct,
                           total_classes=total_classes)


@main.route('/teacher/course/<int:course_id>/export')
@login_required
@role_required('teacher')
def export_attendance(course_id):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{course.code} Attendance"

        # Header
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        headers = ['Roll No', 'Student Name', 'Total Classes', 'Present', 'Absent', 'Percentage', 'Risk Level']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        students = course.get_enrolled_students()
        for row, s in enumerate(students, 2):
            att_q = Attendance.query.filter_by(student_id=s.id, course_id=course_id)
            total = att_q.count()
            present = att_q.filter(Attendance.status.in_(['present', 'late'])).count()
            absent = total - present
            pct = round((present / total * 100), 1) if total > 0 else 0
            risk, _ = s.get_risk_level(course_id)

            ws.cell(row=row, column=1, value=s.roll_no or 'N/A')
            ws.cell(row=row, column=2, value=s.username)
            ws.cell(row=row, column=3, value=total)
            ws.cell(row=row, column=4, value=present)
            ws.cell(row=row, column=5, value=absent)
            ws.cell(row=row, column=6, value=f"{pct}%")
            cell = ws.cell(row=row, column=7, value=risk)

            # Color code by risk
            colors = {'SAFE': 'd4edda', 'CAUTION': 'fff3cd', 'WARNING': 'fde5d4', 'CRITICAL': 'f8d7da'}
            cell.fill = PatternFill(start_color=colors.get(risk, 'ffffff'),
                                    end_color=colors.get(risk, 'ffffff'), fill_type='solid')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'{course.code}_attendance.xlsx')
    except ImportError:
        flash('openpyxl not installed. Cannot export Excel.', 'danger')
        return redirect(url_for('main.course_analytics', course_id=course_id))


# ─── QR Code Routes ────────────────────────────────────────────
@main.route('/teacher/course/<int:course_id>/qr')
@login_required
@role_required('teacher')
def qr_display(course_id):
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    from app.utils.qr_handler import generate_qr_token, generate_qr_image
    token = generate_qr_token(course_id, current_user.id)
    qr_img = generate_qr_image(token, course.name)

    return render_template('qr_display.html', course=course, token=token,
                           qr_img=qr_img, expiry_minutes=5)


@main.route('/scan-qr', methods=['GET', 'POST'])
@login_required
@role_required('student')
def scan_qr():
    if request.method == 'POST':
        token = request.form.get('token', '').strip()
        from app.utils.qr_handler import validate_qr_token
        payload, error = validate_qr_token(token)
        if error:
            flash(error, 'danger')
            return render_template('scan_qr.html')

        course_id = payload['course_id']
        # Check if student is enrolled
        enrollment = Enrollment.query.filter_by(
            student_id=current_user.id, course_id=course_id).first()
        if not enrollment:
            flash('You are not enrolled in this course.', 'danger')
            return render_template('scan_qr.html')

        today = date.today()
        existing = Attendance.query.filter_by(
            student_id=current_user.id, course_id=course_id, date=today).first()
        if existing:
            flash('Attendance already marked for today.', 'info')
        else:
            db.session.add(Attendance(
                student_id=current_user.id, course_id=course_id,
                date=today, status='present'
            ))
            db.session.commit()
            flash('✅ Attendance marked successfully!', 'success')

        return redirect(url_for('main.student_dashboard'))

    return render_template('scan_qr.html')


@main.route('/api/qr/scan', methods=['POST'])
@login_required
def api_qr_scan():
    """API endpoint for QR scanning from camera"""
    data = request.get_json()
    token = data.get('token', '')

    from app.utils.qr_handler import validate_qr_token
    payload, error = validate_qr_token(token.replace('ATTENDANCE:', ''))
    if error:
        return jsonify({'success': False, 'message': error})

    course_id = payload['course_id']
    if current_user.role == 'student':
        enrollment = Enrollment.query.filter_by(
            student_id=current_user.id, course_id=course_id).first()
        if not enrollment:
            return jsonify({'success': False, 'message': 'Not enrolled in this course'})

        today = date.today()
        existing = Attendance.query.filter_by(
            student_id=current_user.id, course_id=course_id, date=today).first()
        if existing:
            return jsonify({'success': False, 'message': 'Attendance already marked'})

        db.session.add(Attendance(
            student_id=current_user.id, course_id=course_id,
            date=today, status='present'
        ))
        db.session.commit()

        # Notify teacher via WebSocket
        course = Course.query.get(course_id)
        socketio.emit('student_scanned', {
            'student_name': current_user.username,
            'roll_no': current_user.roll_no,
            'course_id': course_id,
            'time': datetime.now().strftime('%H:%M:%S')
        }, room=f'course_{course_id}')

        return jsonify({'success': True, 'message': 'Attendance marked!'})

    return jsonify({'success': False, 'message': 'Invalid role'})


# ─── Alert Routes ──────────────────────────────────────────────
@main.route('/teacher/alert/<int:student_id>/<int:course_id>', methods=['POST'])
@login_required
@role_required('teacher')
def send_alert(student_id, course_id):
    student = User.query.get_or_404(student_id)
    course = Course.query.get_or_404(course_id)
    pct = student.get_attendance_percentage(course_id)
    _, classes_needed = student.get_can_miss(course_id)

    results = []
    if student.parent_phone:
        from app.utils.whatsapp import send_low_attendance_alert, send_critical_alert
        if pct < 65:
            ok, msg = send_critical_alert(student, course, pct, classes_needed)
        else:
            ok, msg = send_low_attendance_alert(student, course, pct)
        results.append(f"WhatsApp: {'✅ Sent' if ok else f'❌ {msg}'}")

    if student.parent_email:
        from app.utils.email_handler import send_low_attendance_email
        ok, msg = send_low_attendance_email(student, course, pct)
        results.append(f"Email: {'✅ Sent' if ok else f'❌ {msg}'}")

    if not results:
        flash('No parent contact info found.', 'warning')
    else:
        flash(' | '.join(results), 'info')

    return redirect(url_for('main.course_analytics', course_id=course_id))


# ─── Student Routes ────────────────────────────────────────────
@main.route('/student')
@login_required
@role_required('student')
def student_dashboard():
    enrollments = Enrollment.query.filter_by(student_id=current_user.id).all()
    course_ids = [e.course_id for e in enrollments]
    courses = Course.query.filter(Course.id.in_(course_ids)).all()

    course_stats = []
    for c in courses:
        pct = current_user.get_attendance_percentage(c.id)
        risk, color = current_user.get_risk_level(c.id)
        can_miss, classes_needed = current_user.get_can_miss(c.id)
        att_q = Attendance.query.filter_by(student_id=current_user.id, course_id=c.id)
        total = att_q.count()
        present = att_q.filter(Attendance.status.in_(['present', 'late'])).count()
        course_stats.append({
            'course': c, 'pct': pct, 'risk': risk, 'color': color,
            'can_miss': can_miss, 'classes_needed': classes_needed,
            'total': total, 'present': present
        })

    overall_pct = sum(x['pct'] for x in course_stats) / len(course_stats) if course_stats else 0
    overall_risk, overall_color = current_user.get_risk_level()

    # Recent attendance (last 7 days)
    week_ago = date.today() - timedelta(days=7)
    recent_att = Attendance.query.filter(
        Attendance.student_id == current_user.id,
        Attendance.date >= week_ago
    ).order_by(Attendance.date.desc()).all()

    return render_template('student_dashboard.html',
                           course_stats=course_stats, overall_pct=overall_pct,
                           overall_risk=overall_risk, overall_color=overall_color,
                           recent_att=recent_att)


# ─── WebSocket Events ──────────────────────────────────────────
@socketio.on('join_course')
def on_join_course(data):
    from flask_socketio import join_room
    room = f"course_{data['course_id']}"
    join_room(room)


@socketio.on('leave_course')
def on_leave_course(data):
    from flask_socketio import leave_room
    room = f"course_{data['course_id']}"
    leave_room(room)


# ─── API Endpoints ─────────────────────────────────────────────
@main.route('/api/analytics/overview')
@login_required
@role_required('admin')
def api_analytics_overview():
    """Return system-wide analytics as JSON for charts"""
    students = User.query.filter_by(role='student').all()
    risk_counts = {'SAFE': 0, 'CAUTION': 0, 'WARNING': 0, 'CRITICAL': 0}
    pct_distribution = []

    for s in students:
        pct = s.get_attendance_percentage()
        risk, _ = s.get_risk_level()
        risk_counts[risk] += 1
        pct_distribution.append(pct)

    return jsonify({
        'risk_counts': risk_counts,
        'avg_attendance': sum(pct_distribution) / len(pct_distribution) if pct_distribution else 0,
        'total_students': len(students)
    })


@main.route('/teacher/course/<int:course_id>/student/<int:student_id>/prediction')
@login_required
@role_required('teacher')
def student_prediction(course_id, student_id):
    """Show ML attendance prediction for a specific student."""
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    student = User.query.get_or_404(student_id)

    from app.utils.analytics import get_attendance_trend, predict_future_attendance, get_weekly_pattern
    trend = get_attendance_trend(student_id, course_id, db, Attendance)
    prediction = predict_future_attendance(student_id, course_id, db, Attendance)
    weekly = get_weekly_pattern(student_id, course_id, db, Attendance)

    return render_template('student_prediction.html',
                           course=course, student=student,
                           trend=json.dumps(trend),
                           prediction=prediction,
                           weekly=json.dumps(weekly))


@main.route('/api/student/<int:student_id>/trend/<int:course_id>')
@login_required
def api_student_trend(student_id, course_id):
    """Return attendance trend data as JSON for Chart.js."""
    # Students can only see their own data
    if current_user.role == 'student' and current_user.id != student_id:
        return jsonify({'error': 'Access denied'}), 403

    from app.utils.analytics import get_attendance_trend
    trend = get_attendance_trend(student_id, course_id, db, Attendance)
    return jsonify(trend)


@main.route('/teacher/course/<int:course_id>/live')
@login_required
@role_required('teacher')
def live_dashboard(course_id):
    """Real-time live attendance dashboard via WebSockets."""
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.teacher_dashboard'))

    enrolled_ids = [e.student_id for e in course.enrollments]
    students = User.query.filter(User.id.in_(enrolled_ids)).order_by(User.roll_no).all()
    return render_template('live_dashboard.html', course=course, students=students)
